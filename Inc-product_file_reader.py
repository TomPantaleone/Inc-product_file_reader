#! Python 3.0
#! Inc-product_file_reader.py
#! File that reads a .xls and sorts info in differents lists.

import ast
import os
import shutil
import sys
from os import path
from tkinter import CENTER

import openpyxl
import pyexcel as p
import PySimpleGUI as sg
from openpyxl.styles.borders import Border, Side

pathfolder = (os.path.dirname(__file__))
filefolder = (pathfolder + "\\archivos\\")
workdir = os.chdir(filefolder)
fileinc = os.listdir()[0]
fileinc_ext = os.path.splitext(fileinc)

# Borrar arbol de directorio de la carpeta "pedidos"
for files in os.listdir(pathfolder + "\\pedidos\\"):
    shutil.rmtree(pathfolder + "//pedidos//" + files)

# Archivos donde van a estar los pedidos
# Si el archivo es .xls, guardarlo como .xlsx y abrirlo. Si no lo es, abrirlo
if fileinc_ext[1] == ".xls":
    fileinc_ext2 = fileinc_ext[0] + "_2.xlsx"
    p.save_book_as(file_name=fileinc, dest_file_name=fileinc_ext2)
    wb = openpyxl.load_workbook(fileinc_ext2)
    sheets = wb.sheetnames
    ws = wb[sheets[0]]
else:
    wb = openpyxl.load_workbook(fileinc)
    sheets = wb.sheetnames
    ws = wb[sheets[0]]

# Max cell row + 5 crea un borde, para que pueda leer todo el archivo correctamente. Guarda .xlsx y vuelve a abrir
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

ws.cell(row= ws.max_row+5, column=1).border = thin_border
wb.save("pedidos_generados.xlsx")
wb.close()

wb = openpyxl.load_workbook("pedidos_generados.xlsx")
sheets = wb.sheetnames
ws = wb[sheets[0]]

# Abre el xlsx de pedidos_base
wb_pedidos = openpyxl.load_workbook(pathfolder + "//listados//pedidos.xlsx")
ws_pedidos = wb_pedidos.active

# Abre y crea los diccionarios y las listas desde archivos .txt
file = open(pathfolder + "\listados\productos_kilos.txt", "r")
contents = file.read()
productos = ast.literal_eval(contents)
file.close()

file = open(pathfolder + "\listados\productos_listado.txt", "r")
contents = file.read()
productos_ifco = ast.literal_eval(contents)
file.close()

file = open(pathfolder + "\listados\productos_listado.txt", "r")
contents = file.read()
productos_aca = ast.literal_eval(contents)
file.close()

file = open(pathfolder + "\listados\sucursales_aca.txt", "r")
contents = file.read()
sucursales_aca = ast.literal_eval(contents)
file.close()

file = open(pathfolder + "\listados\sucursales_locacion.txt", "r")
contents = file.read()
sucursales_locacion = ast.literal_eval(contents)
file.close()

# Ventana 
layout = [[sg.Text('Elejir que productos NO se van a contabilizar')],      
            [sg.Checkbox("Acelga", key="ACELGA X PAQUETE      ", size =(14,1)), sg.Checkbox("Puerro", key="PUERRO X ATADO      ", size =(14,1)), sg.Checkbox("Lechuga Mantecosa", key="LECHUGA MANTECOSA X KG.      ", size =(14,1))], 
            [sg.Checkbox("Radicheta", key="RADICHETA X ATADO      ", size =(14,1)), sg.Checkbox("Remolacha", key="REMOLACHA X KG.      ", size =(14,1)), sg.Checkbox("Cebolla de verdeo", key="CEBOLLA DE VERDEO X ATADO      ", size =(14,1))],
            [sg.Checkbox("Lechuga Francesa", key="LECHUGA FRANCESA X KG.      ", size =(14,1)), sg.Checkbox("Albahaca", key="ALBAHACA X ATADO      ", size =(14,1)), sg.Checkbox("Apio", key="APIO X KG.      ", size =(14,1))],
            [sg.Checkbox("Rabanito", key="RABANITO X ATADO.      ", size =(14,1)), sg.Checkbox("Coliflor", key="COLIFLOR x kg.      ", size =(14,1)), sg.Checkbox("Akusay", key="AKUSAY X KG      ", size =(14,1))],
            [sg.Checkbox("Lechuga morada", key="LECHUGA MORADA X KG.      ", size =(14,1)),  sg.Checkbox("Brocoli", key="BROCOLI x kg.      ", size =(14,1)), sg.Checkbox("Repollo colorado", key="REPOLLO COLORADO X KG.      ", size =(14,1))],
            [sg.Checkbox("Lechuga criolla", key="LECHUGA CRIOLLA x kg.      ", size =(14,1)), sg.Checkbox("Escarola ancha", key="LECHUGA ESCAROLA X KG      ", size =(14,1)), sg.Checkbox("Repollo blanco", key="REPOLLO BLANCO x kg.-      ", size =(14,1))],
            [sg.Checkbox("Espinaca", key="ESPINACA X ATADO      ", size =(14,1)), sg.Checkbox("Hinojo", key="HINOJO X KG.      ", size =(14,1)), sg.Checkbox("Rucula", key="RUCULA X ATADO      ", size =(14,1))],
            [sg.Checkbox("Perejil", key="PEREJIL X ATADO      ", size =(14,1)), sg.Checkbox("Lechuga Capuchina", key="LECHUGA CAPUCHINA X KG      ", size =(14,1))],
            [sg.Text("-"*90, size=(60,1), justification= CENTER)],
            [sg.Text("Seleccionar si quiere imprimir los pedidos")],
            [sg.Checkbox("Imprimir Pedidos", key="Print")],
            [sg.Submit("Aceptar"), sg.Cancel("Salir")]]      

window = sg.Window('Automatizador de pedidos', layout)

# Variable para imprimir archivos
imprimir_pedidos = False

while True:
    event, values = window.read() 
    if event == "Aceptar":
        for x,y in values.items():
            if y == True:
                if x in productos_aca.keys():
                    productos_aca.pop(x)
                    productos_ifco.pop(x)
            if x == "Print" and y == True:
                imprimir_pedidos = True
        break
    elif event == sg.WIN_CLOSED or event == "Salir":
        sys.exit()
window.close()

# Crear carpetas con fecha
for cell in range (2, ws.max_row-1):
    date_cell = ws["E" + str(cell)].value
    date = ""
    if date_cell == None:
        break
    elif date_cell != date:
        if path.exists(f"{pathfolder}//pedidos//{date_cell}"):
            continue
        else:
            os.mkdir(f"{pathfolder}//pedidos//{date_cell}")


# Sacando datos para imprimir
suc_value = 0
oc_value = 0
for cell in range (2, ws.max_row-1):
    suc_cell = ws["A" + str(cell)].value                # Suc data
    oc_cell = ws["B" + str(cell)].value               # OC data
    prod_cell = ws["H" + str(cell)].value             # Producto data
    cant_cell = ws["I" + str(cell)].value             # Cantidad data
    date_cell = ws["E" + str(cell)].value               # Fecha data

    # Chequea si la sucursal tiene productos en ifco o aca y los pone en distintas listas
    if suc_cell == None:
        continue
    elif oc_cell != oc_value or suc_cell != suc_value:
        bultos = 0  

        # Le pone la variable "sucursal_locacion" a los .xslx, para que sepa el nombre de la sucursal.
        ws_pedidos["D2"] = suc_cell
        for x in sucursales_locacion.items():
            if suc_cell == x[0]:
                ws_pedidos["F1"] = x[1]
        ws_pedidos["D1"] = oc_cell

        # Saca las cantidades finales (Cajones) - Solo
        if prod_cell in productos_aca.keys() and productos_ifco.keys():
            prod_div = int((cant_cell / productos.get(prod_cell)))
            bultos += prod_div
        # Saca cantidad (Archivo) de todos los productos
            for numero_pedido in range (1, ws_pedidos.max_row-1):
                prod_pedidos = ws_pedidos["C" + str(numero_pedido)].value
                if prod_pedidos == prod_cell:
                    ws_pedidos["B" + str(numero_pedido)] = cant_cell
                    ws_pedidos["F" + str(numero_pedido)] = prod_div

    else:
        if prod_cell in productos_aca.keys() and productos_ifco.keys():
            prod_div = int((cant_cell / productos.get(prod_cell)))
            bultos += prod_div
            for numero_pedido in range (1, ws_pedidos.max_row-1):
                prod_pedidos = ws_pedidos["C" + str(numero_pedido)].value
                if prod_pedidos == prod_cell:
                    ws_pedidos["B" + str(numero_pedido)] = cant_cell
                    ws_pedidos["F" + str(numero_pedido)] = prod_div

    if oc_cell != ws["B" + str(cell+1)].value or suc_cell != ws["A" + str(cell+1)].value:
        ws_pedidos["F31"] = bultos
        if bultos == 0:
            continue
        wb_pedidos.save(f"{pathfolder}//pedidos//{date_cell}//sucursal_{str(suc_cell)}_orden_{str(oc_cell)}.xlsx")

    oc_value = oc_cell
    suc_value = suc_cell

#wb_cantidad = openpyxl.load_workbook(f"{pathfolder}//listados//pedidos.xlsx")
#ws_cantidad = wb_cantidad.active
#ws_cantidad["C1"].value = ""
#ws_cantidad["C2"].value = ""
#for cantidad_pedidos in range (1, ws_cantidad.max_row-1):
#    nomb_pedidos = ws_cantidad["C" + str(cantidad_pedidos)].value
#    for x,y in productos_aca.items():
#        if x == nomb_pedidos and y != 0:
#            ws_cantidad["F" + str(cantidad_pedidos)] = y
#    for x,y in productos_ifco.items():
#        if x == nomb_pedidos and y != 0:
#            ws_cantidad["G" + str(cantidad_pedidos)] = y
#wb_cantidad.save(f"{pathfolder}//pedidos//{date_cell}//cantidad.xlsx")
#wb_cantidad.close()

# Saca las cantidades totales de cada dia.
for folders in os.listdir(f"{pathfolder}//pedidos"):
    for files in os.listdir(f"{pathfolder}//pedidos//{folders}"):
#        print(files)
        wb_sumatoria = openpyxl.load_workbook(f"{pathfolder}//pedidos//{folders}//{files}")
        ws_sumatoria = wb_sumatoria.active
        # Verifica la sucursal si es "aca" o "ifco"
        suc_cell = ws_sumatoria["D2"].value
        if suc_cell in sucursales_aca:
            list_prod = productos_aca
#            print(f"{list_prod} aca")
        else:
            list_prod = productos_ifco
#            print(f"{list_prod} ifco")
        # Saca los valores de cada archivo y se lo asigna a la lista correspondiente ("Aca" o "ifco")
        for cell in range(2,ws_sumatoria.max_row-1):
            prod_cell = ws_sumatoria["C" + str(cell)].value             # Producto data (archivo final)
            cant_cell = ws_sumatoria["F" + str(cell)].value             # Cantidad data (archivo final)
            if prod_cell in productos_aca.keys() and productos_ifco.keys():
                if cant_cell == None:
                    continue
                else:
                    list_prod[prod_cell] += cant_cell

    wb_cantidad = openpyxl.load_workbook(f"{pathfolder}//listados//pedidos.xlsx")
    ws_cantidad = wb_cantidad.active
    ws_cantidad["C1"].value = ""
    ws_cantidad["C2"].value = ""
    for cantidad_pedidos in range (1, ws_cantidad.max_row-1):
        nomb_pedidos = ws_cantidad["C" + str(cantidad_pedidos)].value
        for x,y in productos_aca.items():
            if x == nomb_pedidos and y != 0:
                ws_cantidad["F" + str(cantidad_pedidos)] = y
        for x,y in productos_ifco.items():
            if x == nomb_pedidos and y != 0:
                ws_cantidad["G" + str(cantidad_pedidos)] = y
    wb_cantidad.save(f"{pathfolder}\\pedidos\\{folders}\\cantidad.xlsx")
    wb_cantidad.close()

# Imprimir archivos con checkbox
if imprimir_pedidos == True:
    for root, dirs, files in os.walk(f"{pathfolder}\\pedidos"):
        for file_pedido in files:
            os.startfile(f"{root}//{file}", "print")

file.close()
wb_sumatoria.close()
wb_cantidad.close()